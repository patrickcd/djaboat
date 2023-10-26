from collections import namedtuple
from typing import Any
from openpyxl import load_workbook
from click import confirm

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from inventory.models import Category, Item, Maintenance

Job = namedtuple("Job", ['category', 'task', 'items_required'])

class Command(BaseCommand):
    help = "Closes the specified poll for voting"

    @transaction.atomic
    def handle(self, *args, **options):

        wb = load_workbook("./inventory/management/commands/boating.xlsx")

        maintenance_sheet = wb['Maintenance Log']
        current_category: Category|Any = None
        for row in maintenance_sheet.iter_rows(min_row=5):
            job = Job(row[0].value, row[1].value, row[2].value)

            match job:
                case Job(category=None, task=None):
                    continue
                case Job(category=None):
                    m = Maintenance(task=job.task, category=current_category)
                    # if confirm(f"Save '{m.task}' [{current_category.name}]?", default=True):
                    m.save()
                case _:
                    current_category = Category(name=job.category)
                    current_category.save()

